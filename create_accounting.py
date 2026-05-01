#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台灣小公司 Excel 記帳系統生成器
讀取科目表，生成完整的記帳工作簿
支援：日記簿、試算表、資產負債表、損益表、VAT計算
"""

import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
# 設置
# ============================================================================

INPUT_FILE = r"C:\Users\user\Downloads\捷瑞會計系統企業版2025a.csv"
OUTPUT_FILE = r"C:\users\user\claude-practice\nosif_accounting_v2.xlsx"

# 顏色定義
COLOR_HEADER = "D3D3D3"  # 淺灰
COLOR_TITLE = "4472C4"   # 藍色
COLOR_TOTAL = "FFF2CC"   # 淺黃
COLOR_WARNING = "FFC7CE" # 淺紅

# 邊框定義
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# 讀取科目表
# ============================================================================

def read_coa(filename):
    """從 CSV 讀取科目表，回傳清單"""
    coa = []
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            lines = list(reader)

            # CSV 結構：欄位位置
            # 列 12 = 代號, 列 13 = 啟用, 列 14 = 名稱, 列 15 = 借貸

            # 從第 3 行（索引 2）開始讀取數據（跳過標題）
            for row in lines[2:]:
                if len(row) > 15:
                    code = row[12].strip() if len(row) > 12 else ""
                    name = row[14].strip() if len(row) > 14 else ""
                    direction = row[15].strip() if len(row) > 15 else ""

                    # 驗證：code 必須是數字，name 不能空
                    if code and code.isdigit() and name and len(name) > 0:
                        coa.append({
                            'code': code,
                            'name': name,
                            'direction': direction if direction in ['借', '貸'] else '借'
                        })
    except Exception as e:
        print(f"讀取 CSV 錯誤: {e}")

    return coa

def classify_coa(code):
    """根據科目代號分類為大類（資產/負債/權益等）"""
    first_digit = int(code[0])
    if first_digit == 1:
        return "資產"
    elif first_digit == 2:
        return "負債"
    elif first_digit == 3:
        return "權益"
    elif first_digit == 4:
        return "收入"
    elif first_digit == 5:
        return "銷貨成本"
    elif first_digit == 6:
        return "費用"
    elif first_digit == 7:
        return "非營業收入"
    elif first_digit == 8:
        return "非營業費用"
    elif first_digit == 9:
        return "所得稅"
    else:
        return "其他"

def classify_bs(code):
    """根據科目代號推薦BS分類（用戶可在Excel中覆蓋）"""
    prefix = code[:2]

    if prefix in ['11', '12', '13']:
        return "流動資產"
    elif int(code[0]) == 1:
        return "非流動資產"
    elif prefix in ['21', '22']:
        return "流動負債"
    elif int(code[0]) == 2:
        return "非流動負債"
    elif int(code[0]) == 3:
        return "股東權益"
    else:
        return ""

# ============================================================================
# 建立工作簿
# ============================================================================

def create_accounting_workbook(coa):
    """建立完整的 Excel 工作簿"""
    wb = Workbook()
    wb.remove(wb.active)  # 移除預設 sheet

    # 建立工作表
    ws_guide = wb.create_sheet("說明", 0)
    ws_coa = wb.create_sheet("科目表", 1)
    ws_journal = wb.create_sheet("日記簿", 2)
    ws_tb = wb.create_sheet("試算表", 3)
    ws_bs = wb.create_sheet("資產負債表", 4)
    ws_is = wb.create_sheet("損益表", 5)
    ws_vat = wb.create_sheet("VAT計算", 6)

    # 填充各工作表
    setup_guide(ws_guide)
    setup_coa(ws_coa, coa)
    setup_journal(ws_journal, coa)
    setup_trial_balance(ws_tb, coa)
    setup_balance_sheet(ws_bs, coa)
    setup_income_statement(ws_is, coa)
    setup_vat(ws_vat, coa)

    return wb

# ============================================================================
# Sheet 1: 說明
# ============================================================================

def setup_guide(ws):
    """設置說明頁"""
    ws.column_dimensions['A'].width = 60

    row = 1
    # 標題
    cell = ws[f'A{row}']
    cell.value = "台灣小公司 Excel 記帳系統"
    cell.font = Font(size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=COLOR_TITLE, end_color=COLOR_TITLE, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f'A{row}:D{row}')

    row += 2
    ws[f'A{row}'].value = "使用說明"
    ws[f'A{row}'].font = Font(size=12, bold=True)

    row += 1
    guide_text = [
        "1. 在「日記簿」頁面輸入所有傳票記錄",
        "2. 科目表可在「科目表」頁面查看和修改",
        "3. 「試算表」會自動計算每個科目的借貸合計和餘額",
        "4. 「資產負債表」和「損益表」會自動根據試算表生成",
        "5. 「VAT計算」頁面顯示進項稅和銷項稅",
        "",
        "日記簿輸入欄位：",
        "- 傳票號：唯一編號（如 2401001）",
        "- 日期：YYYY/MM/DD 格式",
        "- 摘要：交易說明",
        "- 科目代號：必須是科目表中的代號",
        "- 借方金額：此科目貸方金額為 0",
        "- 貸方金額：此科目借方金額為 0",
        "",
        "提醒：借方 = 貸方（雙重記帳原則）"
    ]

    for text in guide_text:
        ws[f'A{row}'].value = text
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 20
        row += 1

# ============================================================================
# Sheet 2: 科目表
# ============================================================================

def setup_coa(ws, coa):
    """設置科目表"""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # 標題行
    headers = ["科目代號", "科目名稱", "借/貸", "科目類別", "BS分類"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_TITLE, end_color=COLOR_TITLE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 科目資料
    for row_idx, item in enumerate(coa, 2):
        ws.cell(row_idx, 1).value = item['code']
        ws.cell(row_idx, 2).value = item['name']
        ws.cell(row_idx, 3).value = item['direction']
        ws.cell(row_idx, 4).value = classify_coa(item['code'])
        ws.cell(row_idx, 5).value = classify_bs(item['code'])

        for col in range(1, 6):
            cell = ws.cell(row_idx, col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

# ============================================================================
# Sheet 3: 日記簿
# ============================================================================

def setup_journal(ws, coa):
    """設置日記簿"""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    # 標題行
    headers = ["傳票號", "日期", "摘要", "科目代號", "科目名稱", "借方金額", "貸方金額"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_TITLE, end_color=COLOR_TITLE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 示例行（使用者可刪除）
    ws[f'A2'].value = "2401001"
    ws[f'B2'].value = "2025/01/01"
    ws[f'C2'].value = "公司創立"
    ws[f'D2'].value = "11121"
    ws[f'E2'].value = "=IFERROR(VLOOKUP(D2,科目表!$A:$B,2,FALSE),\"\")"
    ws[f'F2'].value = 1000000
    ws[f'G2'].value = ""

    ws[f'A3'].value = "2401001"
    ws[f'B3'].value = "2025/01/01"
    ws[f'C3'].value = "公司創立"
    ws[f'D3'].value = "3110"
    ws[f'E3'].value = "=IFERROR(VLOOKUP(D3,科目表!$A:$B,2,FALSE),\"\")"
    ws[f'F3'].value = ""
    ws[f'G3'].value = 1000000

    # 設置數字格式、邊框和公式（只到第 100 行避免性能問題）
    for row in range(2, 100):
        # E列：科目名稱自動代出
        if row > 3:  # 第2、3行已有公式
            ws.cell(row, 5).value = f"=IFERROR(VLOOKUP(D{row},科目表!$A:$B,2,FALSE),\"\")"

        ws.cell(row, 6).number_format = '#,##0'
        ws.cell(row, 7).number_format = '#,##0'
        ws.cell(row, 6).border = THIN_BORDER
        ws.cell(row, 7).border = THIN_BORDER
        ws.cell(row, 6).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 7).alignment = Alignment(horizontal="right", vertical="center")

# ============================================================================
# Sheet 4: 試算表
# ============================================================================

def setup_trial_balance(ws, coa):
    """設置試算表"""
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # 標題行
    headers = ["科目代號", "科目名稱", "借方合計", "貸方合計", "餘額", "科目類別", "BS分類"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_TITLE, end_color=COLOR_TITLE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 科目行與公式
    for row_idx, item in enumerate(coa, 2):
        code = item['code']
        name = item['name']
        direction = item['direction']
        category = classify_coa(code)
        bs_category = classify_bs(code)

        ws.cell(row_idx, 1).value = code
        ws.cell(row_idx, 2).value = name

        # 借方合計公式（限制到第 100 行避免兼容性問題）
        ws.cell(row_idx, 3).value = f"=SUMIFS(日記簿!$F$2:$F$100,日記簿!$D$2:$D$100,A{row_idx})"
        ws.cell(row_idx, 3).number_format = '#,##0'

        # 貸方合計公式
        ws.cell(row_idx, 4).value = f"=SUMIFS(日記簿!$G$2:$G$100,日記簿!$D$2:$D$100,A{row_idx})"
        ws.cell(row_idx, 4).number_format = '#,##0'

        # 餘額公式（根據借/貸方向）
        if direction == "借":
            ws.cell(row_idx, 5).value = f"=C{row_idx}-D{row_idx}"
        else:
            ws.cell(row_idx, 5).value = f"=D{row_idx}-C{row_idx}"
        ws.cell(row_idx, 5).number_format = '#,##0'

        # 科目類別（用於 IS 計算）
        ws.cell(row_idx, 6).value = category

        # BS分類（用於資產負債表計算）
        ws.cell(row_idx, 7).value = bs_category

        # 格式設置
        for col in range(1, 8):
            cell = ws.cell(row_idx, col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col >= 3 else "left", vertical="center")

    # 合計行
    total_row = len(coa) + 2
    ws.cell(total_row, 2).value = "合計"
    ws.cell(total_row, 2).font = Font(bold=True)
    ws.cell(total_row, 2).fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")

    ws.cell(total_row, 3).value = f"=SUM(C2:C{total_row-1})"
    ws.cell(total_row, 3).font = Font(bold=True)
    ws.cell(total_row, 3).fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws.cell(total_row, 3).number_format = '#,##0'
    ws.cell(total_row, 3).border = THIN_BORDER

    ws.cell(total_row, 4).value = f"=SUM(D2:D{total_row-1})"
    ws.cell(total_row, 4).font = Font(bold=True)
    ws.cell(total_row, 4).fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws.cell(total_row, 4).number_format = '#,##0'
    ws.cell(total_row, 4).border = THIN_BORDER

    # 平衡檢查
    ws.cell(total_row + 2, 1).value = "借貸平衡檢查:"
    ws.cell(total_row + 2, 1).font = Font(bold=True)
    ws.cell(total_row + 2, 3).value = f"=IF(C{total_row}=D{total_row},\"✓ 平衡\",\"✗ 不平衡\")"
    ws.cell(total_row + 2, 3).font = Font(bold=True, color="008000")

# ============================================================================
# Sheet 5: 資產負債表
# ============================================================================

def setup_balance_sheet(ws, coa):
    """設置資產負債表（左右並排）"""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 2
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15

    # 標題
    ws['A1'] = "資產負債表"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"截至 {datetime.now().strftime('%Y年%m月%d日')}"
    ws['A2'].font = Font(size=10)

    row_left = 4
    row_right = 4
    coa_count = len(coa)

    # 幫助函數：添加BS類別的科目明細（左側）
    def add_bs_category_left(category_name):
        nonlocal row_left

        ws[f'A{row_left}'] = category_name
        ws[f'A{row_left}'].font = Font(bold=True, size=11)
        ws[f'A{row_left}'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        row_left += 1

        items = [item for item in coa if classify_bs(item['code']) == category_name]
        subtotal_rows = []

        for item in items:
            code = item['code']
            name = item['name']
            ws[f'A{row_left}'] = f"  {code} {name}"
            ws[f'A{row_left}'].font = Font(size=9)
            ws[f'B{row_left}'] = f"=IFERROR(VLOOKUP(\"{code}\",試算表!$A:$E,5,FALSE),0)"
            ws[f'B{row_left}'].number_format = '#,##0'
            ws[f'B{row_left}'].font = Font(size=9)
            subtotal_rows.append(row_left)
            row_left += 1

        if subtotal_rows:
            start_row = subtotal_rows[0]
            end_row = subtotal_rows[-1]
            ws[f'A{row_left}'] = f"{category_name}合計"
            ws[f'A{row_left}'].font = Font(bold=True, size=10)
            ws[f'B{row_left}'] = f"=SUM(B{start_row}:B{end_row})"
            ws[f'B{row_left}'].number_format = '#,##0'
            ws[f'B{row_left}'].font = Font(bold=True, size=10)
            subtotal_row = row_left
        else:
            subtotal_row = row_left
            ws[f'B{row_left}'] = 0
            ws[f'B{row_left}'].number_format = '#,##0'

        row_left += 1
        return subtotal_row

    # 幫助函數：添加BS類別的科目明細（右側）
    def add_bs_category_right(category_name):
        nonlocal row_right

        ws[f'D{row_right}'] = category_name
        ws[f'D{row_right}'].font = Font(bold=True, size=11)
        ws[f'D{row_right}'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        row_right += 1

        items = [item for item in coa if classify_bs(item['code']) == category_name]
        subtotal_rows = []

        for item in items:
            code = item['code']
            name = item['name']
            ws[f'D{row_right}'] = f"  {code} {name}"
            ws[f'D{row_right}'].font = Font(size=9)
            ws[f'E{row_right}'] = f"=IFERROR(VLOOKUP(\"{code}\",試算表!$A:$E,5,FALSE),0)"
            ws[f'E{row_right}'].number_format = '#,##0'
            ws[f'E{row_right}'].font = Font(size=9)
            subtotal_rows.append(row_right)
            row_right += 1

        if subtotal_rows:
            start_row = subtotal_rows[0]
            end_row = subtotal_rows[-1]
            ws[f'D{row_right}'] = f"{category_name}合計"
            ws[f'D{row_right}'].font = Font(bold=True, size=10)
            ws[f'E{row_right}'] = f"=SUM(E{start_row}:E{end_row})"
            ws[f'E{row_right}'].number_format = '#,##0'
            ws[f'E{row_right}'].font = Font(bold=True, size=10)
            subtotal_row = row_right
        else:
            subtotal_row = row_right
            ws[f'E{row_right}'] = 0
            ws[f'E{row_right}'].number_format = '#,##0'

        row_right += 1
        return subtotal_row

    # 資產側標題
    ws[f'A{row_left}'] = "資產"
    ws[f'A{row_left}'].font = Font(bold=True, size=12)
    ws[f'A{row_left}'].fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    row_left += 1

    ca_total = add_bs_category_left("流動資產")
    nca_total = add_bs_category_left("非流動資產")

    ws[f'A{row_left}'] = "資產總計"
    ws[f'A{row_left}'].font = Font(bold=True, size=11)
    ws[f'A{row_left}'].fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws[f'B{row_left}'] = f"=B{ca_total}+B{nca_total}"
    ws[f'B{row_left}'].font = Font(bold=True, size=11)
    ws[f'B{row_left}'].fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws[f'B{row_left}'].number_format = '#,##0'
    total_assets_row = row_left
    row_left += 1

    # 負債及權益側標題
    ws[f'D{row_right}'] = "負債及權益"
    ws[f'D{row_right}'].font = Font(bold=True, size=12)
    ws[f'D{row_right}'].fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    row_right += 1

    cl_total = add_bs_category_right("流動負債")
    ncl_total = add_bs_category_right("非流動負債")
    eq_total = add_bs_category_right("股東權益")

    ws[f'D{row_right}'] = "負債及權益總計"
    ws[f'D{row_right}'].font = Font(bold=True, size=11)
    ws[f'D{row_right}'].fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws[f'E{row_right}'] = f"=E{cl_total}+E{ncl_total}+E{eq_total}"
    ws[f'E{row_right}'].font = Font(bold=True, size=11)
    ws[f'E{row_right}'].fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws[f'E{row_right}'].number_format = '#,##0'
    total_le_row = row_right
    row_right += 1

    # 驗證（放在最下面）
    max_row = max(row_left, row_right) + 2
    ws[f'A{max_row}'] = "驗證："
    ws[f'B{max_row}'] = f"=IF(B{total_assets_row}=E{total_le_row},\"✓ 平衡\",\"✗ 不平衡\")"
    ws[f'B{max_row}'].font = Font(bold=True, size=10, color="008000")

# ============================================================================
# Sheet 6: 損益表
# ============================================================================

def setup_income_statement(ws, coa):
    """設置損益表（含科目細項）"""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    # 標題
    ws['A1'] = "損益表"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"自 {datetime.now().strftime('%Y年%m月')} 起"
    ws['A2'].font = Font(size=10)

    row = 4
    coa_count = len(coa)

    # 幫助函數：添加科目類別的細項
    def add_category_items(category_name, display_name):
        nonlocal row
        # 標題行
        ws[f'A{row}'] = display_name
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'B{row}'] = f"=SUMIF(試算表!$F$2:$F${coa_count+1},\"{category_name}\",試算表!$E$2:$E${coa_count+1})"
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'B{row}'].font = Font(bold=True, size=11)
        title_row = row
        row += 1

        # 查找該類別的所有科目
        category_items = [item for item in coa if classify_coa(item['code']) == category_name]

        # 添加科目細項
        for item in category_items:
            ws[f'A{row}'] = f"  {item['code']} {item['name']}"
            ws[f'A{row}'].font = Font(size=9)
            ws[f'B{row}'] = f"=VLOOKUP({item['code']},試算表!$A:$E,5,FALSE)"
            ws[f'B{row}'].number_format = '#,##0'
            ws[f'B{row}'].font = Font(size=9)
            row += 1

        return title_row

    # 營業收入（4xxx）
    income_row = add_category_items("收入", "營業收入")
    row += 1

    # 銷貨成本（5xxx）
    cogs_row = add_category_items("銷貨成本", "銷貨成本")
    row += 1

    # 營業毛利
    ws[f'A{row}'] = "營業毛利"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"=B{income_row}-B{cogs_row}"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '#,##0'
    gp_row = row
    row += 2

    # 營業費用（6xxx）
    expense_row = add_category_items("費用", "營業費用")
    row += 1

    # 營業淨利
    ws[f'A{row}'] = "營業淨利"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"=B{gp_row}-B{expense_row}"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '#,##0'
    oi_row = row
    row += 2

    # 稅前損益
    ws[f'A{row}'] = "稅前損益"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"=B{oi_row}"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '#,##0'
    pbt_row = row
    row += 1

    # 所得稅費用（9xxx）
    tax_row = add_category_items("所得稅", "所得稅費用")
    row += 1

    # 稅後損益
    ws[f'A{row}'] = "稅後損益"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws[f'B{row}'] = f"=B{pbt_row}-B{tax_row}"
    ws[f'B{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'].fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws[f'B{row}'].number_format = '#,##0'

# ============================================================================
# Sheet 7: VAT 計算
# ============================================================================

def setup_vat(ws, coa):
    """設置 VAT 計算"""
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    # 標題
    ws['A1'] = "增值稅（VAT）計算"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = "台灣營業稅率：5%"
    ws['A2'].font = Font(size=10)

    row = 4

    # 進項稅額
    ws[f'A{row}'] = "進項稅額（科目 1144）"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'] = "=VLOOKUP(\"1144\",試算表!$A:$E,5,FALSE)"
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = Font(bold=True)
    row += 2

    # 銷項稅額
    ws[f'A{row}'] = "銷項稅額（科目 2171）"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'] = "=VLOOKUP(\"2171\",試算表!$A:$E,5,FALSE)"
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = Font(bold=True)
    row += 2

    # 應繳 VAT
    ws[f'A{row}'] = "應繳增值稅"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws[f'B{row}'] = f"=B{row-2}-B{row-4}"
    ws[f'B{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'].fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")
    ws[f'B{row}'].number_format = '#,##0'

    row += 3
    ws[f'A{row}'] = "說明："
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "應繳 VAT > 0：需繳稅"
    row += 1
    ws[f'A{row}'] = "應繳 VAT < 0：可退稅（負數表示待退）"
    row += 1
    ws[f'A{row}'] = "應繳 VAT = 0：不需繳退"

# ============================================================================
# 主程式
# ============================================================================

def main():
    print("開始生成 Excel 記帳工作簿...")

    # 讀取科目表
    print(f"讀取科目表：{INPUT_FILE}")
    coa = read_coa(INPUT_FILE)
    print(f"找到 {len(coa)} 個科目")

    if not coa:
        print("警告：未找到科目資料，使用預設科目")
        coa = [
            # 資產
            {'code': '1111', 'name': '零用金', 'direction': '借'},
            {'code': '1112', 'name': '銀行存款', 'direction': '借'},
            {'code': '11121', 'name': '銀行存款-國泰', 'direction': '借'},
            {'code': '11122', 'name': '銀行存款-合庫', 'direction': '借'},
            # 權益
            {'code': '3110', 'name': '實收資本', 'direction': '貸'},
            # 收入
            {'code': '4100', 'name': '銷貨收入', 'direction': '貸'},
            # 費用
            {'code': '6230', 'name': '其他費用', 'direction': '借'},
            # VAT
            {'code': '1144', 'name': '進項稅額', 'direction': '借'},
            {'code': '2171', 'name': '銷項稅額', 'direction': '貸'},
        ]

    # 建立工作簿
    print("建立 Excel 工作簿...")
    wb = create_accounting_workbook(coa)

    # 保存
    print(f"保存文件：{OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)

    print("[OK] 完成!")
    print(f"輸出文件：{OUTPUT_FILE}")
    print("\n使用方式：")
    print("1. 在「日記簿」頁面輸入傳票")
    print("2. 「試算表」會自動計算")
    print("3. 「資產負債表」和「損益表」會自動生成")

if __name__ == "__main__":
    main()
